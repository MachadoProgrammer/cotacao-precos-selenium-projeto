# Tecnologias
import pyperclip
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import openpyxl
import logging as lg
import schedule
from time import sleep
import os

# Firefox -> geckodriver

lg.basicConfig(level=lg.INFO, filename='app.log', format='%(asctime)s %(levelname)s %(message)s')


def rodar_bot_de_pesquisa_por_menor_preco():
    def iniciar_driver(site_url, detach=False, headless=False):
        lg.getLogger(__name__)
        try:
            chorme_options = Options()
            # Standard configuration for chrome
            '''
            --start-maximized # Inicia maximizado
            --lang=pt-BR # Define o idioma de inicialização, # en-US , pt-BR
            --incognito # Usar o modo anônimo
            --no-default-browser-check', # Desabilita a busca pelo browser default
            --window-position=0,0', # Define o posicionamento da janela
            --window-size=800,800 # Define a resolução da janela em largura e altura
            --headless # Roda em segundo plano(com a janela fechada)
            --disable-notifications # Desabilita notificações
            --disable-gpu # Desabilita renderização com GPU
            --block-new-web-contents', # Bloqueia pop-ups
            # para rodar na aws ec2:
            # '--disable-gpu', '--no-sandbox', '--headless', '--disable-dev-sh
            '''

            arguments = [
                '--lang=pt-BR',
                '--window-size=1300,1000',
                '--disable-notifications',
                '--incognito'
            ]
            for argument in arguments:
                chorme_options.add_argument(argument)

            # configuracoes adicionais do options:

            # rodar em segundo plano
            if headless:
                chorme_options.add_argument('--headless')

            # manter janela aberta
            if detach:
                chorme_options.add_experimental_option('detach', True)

            # desabilitar pop-up de navegador controlado por automacao
            chorme_options.add_experimental_option(
                'excludeSwitches', ['enable-automation'])

            # Using experimental settings
            chorme_options.add_experimental_option('prefs', {
                # Alterar o local padrão de ‘download’ de arquivos
                'download.default_directory': 'C:\\Users\\gabri\\PycharmProjects\\Mentorias\\MentoriaProjeto1',
                # notificar o Google chrome sobre essa alteração
                'download.directory_upgrade': True,
                # Desabilitar a confirmação de ‘download’
                'download.prompt_for_download': False,
                # Desabilitar notificações
                'profile.default_content_setting_values.notifications': 2,
                # Permitir multiplos downloads
                'profile.default_content_setting_values.automatic_downloads': 1,

            })

            # Initializing webdriver

            driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=chorme_options)
            driver.get(site_url)

            wait = WebDriverWait(
                driver,
                10,
                poll_frequency=1,
                ignored_exceptions=[
                    NoSuchElementException,
                    ElementNotVisibleException,
                    ElementNotSelectableException,
                    TimeoutException
                ]
            )

            return driver, wait
        except Exception as e:
            lg.error(f'Error occurred while initializng driver: {type(e).__name__} - {e}')
            return None

    def obter_precos():
        driver, wait = iniciar_driver('https://www.kabum.com.br/busca/memoria-16bg')
        preco_kabum = wait.until(
            EC.visibility_of_all_elements_located((By.XPATH, "//span[contains(@class, 'priceCard')]")))
        preco_kabum1 = preco_kabum[0].text.split(' ')[1].replace(',', '.')

        driver.get('https://www.studiopc.com.br/produtos?q=memoria+ram+16gb')
        preco_studio = wait.until(EC.visibility_of_any_elements_located((By.XPATH, '//span[@class="price total"]')))
        preco_studio1 = preco_studio[0].text.split(' ')[1].replace(',', '.')

        return float(preco_kabum1), float(preco_studio1)

    def gerar_planilha_margem_de_lucro():
        preco_site1, preco_site2 = obter_precos()
        custo = 200

        site1 = 'https://www.kabum.com.br/busca/memoria-16bg'
        site2 = 'https://www.studiopc.com.br/produtos?q=memoria+ram+16gb'

        workbook = openpyxl.Workbook()
        del workbook['Sheet']
        workbook.create_sheet('margem_lucro')
        sheet_margem_lucro = workbook['margem_lucro']

        # Passar uma lista, cada item corresponderá a uma coluna
        sheet_margem_lucro.append(['Site', 'custo', 'Preço', 'Lucro'])
        sheet_margem_lucro.append([site1, custo, preco_site1, preco_site1 - custo])
        sheet_margem_lucro.append([site2, custo, preco_site2, preco_site2 - custo])

        workbook.save('margem de lucro.xlsx')

        workbook_margem_lucro = openpyxl.load_workbook('margem de lucro.xlsx')
        sheet_margem_lucro = workbook_margem_lucro['margem_lucro']

        margem_de_lucro = ''
        for row in sheet_margem_lucro.iter_rows(min_row=1):
            # Site, Custo, Preço, Lucro
            margem_de_lucro += f'{row[0].value},{row[1].value},{row[2].value},{row[3].value}'

        with open('margem_lucro.txt', 'w', encoding='utf-8', newline='') as f:
            f.write(margem_de_lucro + os.linesep)

        return margem_de_lucro

    def enviar_margem_de_lucro_no_whatsapp():
        # 3 - enviar em um grupo do whatsapp
        # apertar windows no meu teclado
        margem_de_lucro = gerar_planilha_margem_de_lucro()
        pyautogui.hotkey('win')
        sleep(5)
        # digitar whatsapp
        pyautogui.write('whatsapp')
        sleep(5)
        # apertar enter
        pyautogui.hotkey('enter')
        sleep(5)
        # digitar no campo de pesquisa o nome do grupo
        botao_mais = pyautogui.locateCenterOnScreen('botao_mais.png')
        sleep(2)
        pyautogui.moveTo(botao_mais[0], botao_mais[1], duration=2)
        sleep(2)
        pyautogui.move(0, 50)
        sleep(2)
        pyautogui.click()
        sleep(2)
        pyautogui.write('Grupo T.I.')
        sleep(2)
        # apertar enter
        pyautogui.hotkey('enter')
        sleep(2)
        # digitar o que for necessário
        pyperclip.copy(margem_de_lucro)
        pyautogui.hotkey('ctrl', 'v')
        sleep(2)
        # apertar enter
        pyautogui.hotkey('enter')
        sleep(2)

        enviar_margem_de_lucro_no_whatsapp()


# 4 - Agendar para que o código rode todos os dias as 06:00 da manhã
schedule.every().day.at("00:06").do(rodar_bot_de_pesquisa_por_menor_preco)
rodar_bot_de_pesquisa_por_menor_preco()
while True:
    schedule.run_pending()
    sleep(1)
